"""
Sample JSON File

[
  {
    "quote": "The only limit to our realization of tomorrow is our doubts of today.",
    "date": "2024-08-10"
  },
  {
    "quote": "Imagination is more important than knowledge.",
    "date": "2023-12-15"
  },
  {
    "quote": "In the middle of difficulty lies opportunity.",
    "date": "2022-05-20"
  },
  {
    "quote": "To confine our attention to terrestrial matters would be to limit the human spirit.",
    "date": "2021-11-03"
  },
  {
    "quote": "Somewhere, something incredible is waiting to be known.",
    "date": "2020-07-19"
  }
]

"""

import json
import os
import requests
import random
from datetime import datetime, timedelta
from moviepy import *
from moviepy.video.fx import *
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from PIL import Image
from io import BytesIO

# --- Constants ---
OUTPUT_FOLDER = "generated_shorts"
QUOTES_FILE = os.path.join(OUTPUT_FOLDER, "quotes.json")
DONE_FILE = os.path.join(OUTPUT_FOLDER, "shorts_done.txt")
EXCEL_FILE = os.path.join(OUTPUT_FOLDER, "video_status.xlsx")
NASA_API_KEY = "c9ejGRmHwjE9S0ui6x4Me3nSRsJUAMEeEfhEQARE"  
NASA_API_URL = "https://api.nasa.gov/planetary/apod"

def fetch_nasa_image(date_str):
    """Fetches NASA APOD image for a given date."""
    params = {"api_key": NASA_API_KEY, "date": date_str}
    response = requests.get(NASA_API_URL, params=params)
    if response.status_code == 200:
        data = response.json()
        if data.get("media_type") == "image":
            img_url = data.get("hdurl") or data.get("url")
            img_resp = requests.get(img_url)
            if img_resp.status_code == 200:
                image_path = os.path.join(OUTPUT_FOLDER, f"background_{date_str}.jpg")
                with open(image_path, 'wb') as f:
                    f.write(img_resp.content)
                return image_path
    return None

def get_text_color(image_path):
    """Returns 'white' or 'black' for best contrast with background."""
    img = Image.open(image_path).convert('L').resize((50, 50))
    avg_brightness = sum(img.getdata()) / (50 * 50)
    return "black" if avg_brightness > 128 else "white"

def wrap_text(text, width=40):
    """Wraps text to fit the screen."""
    import textwrap
    return "\n".join(textwrap.wrap(text, width=width))

def get_audio_duration(audio_path):
    """Returns the duration of the audio file in seconds."""
    return AudioFileClip(audio_path).duration

def create_short(quote, audio_path, output_path, date_str):
    """Creates a short video with the given quote and NASA image."""
    duration = 15  # seconds

    # --- Background Image ---
    image_path = fetch_nasa_image(date_str)
    if not image_path:
        print(f"Could not fetch NASA image for {date_str}.")
        return

    # --- Text Styling ---
    text_color = get_text_color(image_path)
    wrapped_quote = wrap_text(quote, width=35)

    # --- Background Clip ---
    bg = ImageClip(image_path).with_duration(duration).resized(height=1920).resized(width=1080)

    # --- Audio ---
    if os.path.exists(audio_path):
        print(f"Audio file {audio_path} does not exist.")
        
        audio_total_duration = get_audio_duration(audio_path)
        if audio_total_duration > duration:
            max_start = audio_total_duration - duration
            start_time = random.uniform(0, max_start)
        else:
            start_time = 0
        audio = AudioFileClip(audio_path).subclipped(start_time, start_time + duration)
        bg = bg.with_audio(audio)

    # --- Text Overlay ---
    txt_clip = (TextClip(
        text=wrapped_quote,
        font_size=55,
        color=text_color,
        size=(1000, None),
        method="caption",
    ).with_duration(duration)
     .with_position(("center", "center"))
     .with_effects([FadeIn(duration=1)]))

    # --- Final Video ---
    video = CompositeVideoClip([bg, txt_clip])
    video.write_videofile(output_path, fps=30, codec="libx264", audio_codec="aac")
    os.remove(image_path)  # Clean up the downloaded image

def setup_excel_file():
    """Creates the Excel file with headers and dropdown if it doesn't exist."""
    if not os.path.exists(EXCEL_FILE):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Video Status"
        sheet["A1"] = "Quote"
        sheet["B1"] = "Output File"
        sheet["C1"] = "Upload Status"

        # Create dropdown for status
        dv = DataValidation(type="list", formula1='"Uploaded,Not Uploaded"', allow_blank=True)
        sheet.add_data_validation(dv)
        dv.add('C2:C1048576')

        workbook.save(EXCEL_FILE)

def update_excel_file(quote, output_filename):
    """Adds a new row to the Excel file for the created video."""
    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active
    sheet.append([quote, output_filename, "Not Uploaded"])
    workbook.save(EXCEL_FILE)

def main():
    """Main function to generate shorts."""
    if not os.path.exists(OUTPUT_FOLDER):
        os.makedirs(OUTPUT_FOLDER)

    setup_excel_file()

    if not os.path.exists(QUOTES_FILE):
        print(f"Error: {QUOTES_FILE} not found.")
        return

    with open(QUOTES_FILE, 'r') as f:
        quotes = json.load(f)

    done_quotes = []
    if os.path.exists(DONE_FILE):
        with open(DONE_FILE, 'r') as f:
            done_quotes = [line.strip() for line in f.readlines()]

    # Ask how many shorts to create if more than 10
    quotes_to_process = [item for item in quotes if item.get("quote") and item.get("quote") not in done_quotes]
    if len(quotes_to_process) > 10:
        try:
            n = int(input(f"There are {len(quotes_to_process)} quotes. How many shorts do you want to create? "))
            quotes_to_process = quotes_to_process[:n]
        except Exception:
            print("Invalid input. Creating 10 shorts by default.")
            quotes_to_process = quotes_to_process[:10]

    for item in quotes_to_process:
        quote = item.get("quote")
        # Use date from item if present, else use today minus random days
        date_str = item.get("date")
        if not date_str:
            random_days = random.randint(0, 3650)
            date_str = (datetime.today() - timedelta(days=random_days)).strftime("%Y-%m-%d")
        print(f"Creating short for: {quote} (Date: {date_str})")
        output_filename = os.path.join(OUTPUT_FOLDER, f"{quote.replace(' ', '_')}_short.mp4")
        create_short(quote, "audio.mp3", output_filename, date_str)
        with open(DONE_FILE, 'a') as f:
            f.write(quote + '\n')
        update_excel_file(quote, output_filename)
        print(f"Finished short: {output_filename}")

if __name__ == "__main__":
    main()